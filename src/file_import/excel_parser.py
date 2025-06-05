"""
Excelファイルパーサー
セキュアで柔軟なExcelファイル解析機能を提供
"""

import logging
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import io

logger = logging.getLogger(__name__)


@dataclass
class ExcelParseResult:
    """Excel解析結果"""
    data: Dict[str, List[Dict[str, Any]]]  # シート名をキーとするデータ
    sheet_names: List[str]
    row_counts: Dict[str, int]
    headers: Dict[str, List[str]]
    has_headers: Dict[str, bool]
    errors: List[str]
    warnings: List[str]


class ExcelParser:
    """
    Excelファイルの安全で柔軟な解析を行うクラス
    """
    
    # サポートするファイル形式
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls', '.xlsm']
    
    # デフォルト設定
    DEFAULT_MAX_ROWS = 50000
    DEFAULT_MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
    DEFAULT_MAX_SHEETS = 20
    
    def __init__(self, 
                 max_rows: int = DEFAULT_MAX_ROWS,
                 max_file_size: int = DEFAULT_MAX_FILE_SIZE,
                 max_sheets: int = DEFAULT_MAX_SHEETS):
        """
        Excelパーサーを初期化
        
        Args:
            max_rows: シートあたりの最大読み込み行数
            max_file_size: 最大ファイルサイズ（バイト）
            max_sheets: 最大シート数
        """
        self.max_rows = max_rows
        self.max_file_size = max_file_size
        self.max_sheets = max_sheets
    
    def parse_file(self, file_path: Union[str, Path], 
                   sheet_names: Optional[List[str]] = None,
                   has_header: Optional[bool] = None) -> ExcelParseResult:
        """
        Excelファイルを解析
        
        Args:
            file_path: Excelファイルのパス
            sheet_names: 読み込むシート名のリスト（Noneの場合は全シート）
            has_header: ヘッダー行の有無（Noneの場合は自動判定）
            
        Returns:
            ExcelParseResult: 解析結果
        """
        file_path = Path(file_path)
        
        # ファイル存在チェック
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # ファイル拡張子チェック
        if file_path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        # ファイルサイズチェック
        if file_path.stat().st_size > self.max_file_size:
            raise ValueError(f"File size exceeds maximum allowed size: {self.max_file_size}")
        
        # ファイル読み込み
        with open(file_path, 'rb') as f:
            content = f.read()
        
        return self.parse_content(content, sheet_names, has_header)
    
    def parse_content(self, content: bytes,
                     sheet_names: Optional[List[str]] = None,
                     has_header: Optional[bool] = None) -> ExcelParseResult:
        """
        Excelコンテンツを解析
        
        Args:
            content: Excelファイルの内容（バイト）
            sheet_names: 読み込むシート名のリスト
            has_header: ヘッダー行の有無
            
        Returns:
            ExcelParseResult: 解析結果
        """
        errors = []
        warnings = []
        
        try:
            # openpyxlを使用してワークブックを読み込み
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            
        except InvalidFileException as e:
            logger.error(f"Invalid Excel file: {e}")
            errors.append(f"Invalid Excel file format: {str(e)}")
            return self._empty_result(errors, warnings)
        
        except Exception as e:
            logger.exception(f"Failed to load Excel file: {e}")
            errors.append(f"Failed to load Excel file: {str(e)}")
            return self._empty_result(errors, warnings)
        
        try:
            # シート名を取得
            available_sheets = workbook.sheetnames
            logger.debug(f"Available sheets: {available_sheets}")
            
            # シート数制限チェック
            if len(available_sheets) > self.max_sheets:
                warnings.append(f"File has {len(available_sheets)} sheets, processing only first {self.max_sheets}")
                available_sheets = available_sheets[:self.max_sheets]
            
            # 読み込み対象シートを決定
            if sheet_names is None:
                target_sheets = available_sheets
            else:
                target_sheets = [sheet for sheet in sheet_names if sheet in available_sheets]
                missing_sheets = [sheet for sheet in sheet_names if sheet not in available_sheets]
                if missing_sheets:
                    warnings.append(f"Sheets not found: {missing_sheets}")
            
            if not target_sheets:
                warnings.append("No valid sheets found to process")
                return self._empty_result(errors, warnings)
            
            # 各シートを解析
            all_data = {}
            all_headers = {}
            all_has_headers = {}
            all_row_counts = {}
            
            for sheet_name in target_sheets:
                try:
                    sheet_data, sheet_headers, sheet_has_header = self._parse_sheet(
                        workbook[sheet_name], has_header
                    )
                    
                    all_data[sheet_name] = sheet_data
                    all_headers[sheet_name] = sheet_headers
                    all_has_headers[sheet_name] = sheet_has_header
                    all_row_counts[sheet_name] = len(sheet_data)
                    
                    # 行数制限チェック
                    if len(sheet_data) > self.max_rows:
                        warnings.append(f"Sheet '{sheet_name}' truncated to {self.max_rows} rows")
                        all_data[sheet_name] = sheet_data[:self.max_rows]
                        all_row_counts[sheet_name] = self.max_rows
                        
                except Exception as e:
                    logger.exception(f"Failed to parse sheet '{sheet_name}': {e}")
                    errors.append(f"Failed to parse sheet '{sheet_name}': {str(e)}")
                    continue
            
            return ExcelParseResult(
                data=all_data,
                sheet_names=list(all_data.keys()),
                row_counts=all_row_counts,
                headers=all_headers,
                has_headers=all_has_headers,
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            logger.exception(f"Failed to process Excel file: {e}")
            errors.append(f"Excel processing failed: {str(e)}")
            return self._empty_result(errors, warnings)
        
        finally:
            if 'workbook' in locals():
                workbook.close()
    
    def _parse_sheet(self, worksheet, has_header: Optional[bool] = None) -> tuple[List[Dict[str, Any]], List[str], bool]:
        """
        ワークシートを解析してデータとヘッダーを抽出
        
        Args:
            worksheet: openpyxlワークシート
            has_header: ヘッダー行の有無
            
        Returns:
            tuple: (データ, ヘッダー, ヘッダー検出結果)
        """
        # セルデータを行として読み込み
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            # 空行でない場合のみ追加
            if any(cell is not None and str(cell).strip() for cell in row):
                rows.append([self._clean_cell_value(cell) for cell in row])
        
        if not rows:
            return [], [], False
        
        # 全行の最大列数を取得
        max_cols = max(len(row) for row in rows) if rows else 0
        
        # 行の長さを統一（短い行は None で埋める）
        normalized_rows = []
        for row in rows:
            normalized_row = row + [None] * (max_cols - len(row))
            normalized_rows.append(normalized_row)
        
        # ヘッダー判定
        if has_header is None:
            has_header = self._detect_header(normalized_rows)
        
        if has_header and len(normalized_rows) > 1:
            header_row = normalized_rows[0]
            data_rows = normalized_rows[1:]
        else:
            # ヘッダーがない場合は自動生成
            if normalized_rows:
                num_cols = len(normalized_rows[0])
                header_row = [f"Column_{i+1}" for i in range(num_cols)]
                data_rows = normalized_rows
            else:
                header_row = []
                data_rows = []
        
        # ヘッダーの重複を解決
        headers = self._resolve_duplicate_headers(header_row)
        
        # データを辞書形式に変換
        data = []
        for row in data_rows:
            # 空行をスキップ
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = value
            
            data.append(row_dict)
        
        return data, headers, has_header
    
    def _detect_header(self, rows: List[List[Any]]) -> bool:
        """
        ヘッダー行の存在を検出
        
        Args:
            rows: 行データ
            
        Returns:
            bool: ヘッダー行が存在するかどうか
        """
        if len(rows) < 2:
            return False
        
        first_row = rows[0]
        second_row = rows[1]
        
        # 1行目に文字列が多く、2行目に数値が多い場合はヘッダーありと判定
        first_row_text_count = 0
        second_row_numeric_count = 0
        
        for cell1, cell2 in zip(first_row, second_row):
            if cell1 is not None and isinstance(cell1, str) and cell1.strip():
                first_row_text_count += 1
            
            if cell2 is not None and isinstance(cell2, (int, float)):
                second_row_numeric_count += 1
        
        # 判定ロジック
        total_cols = len(first_row)
        if total_cols == 0:
            return False
        
        text_ratio = first_row_text_count / total_cols
        numeric_ratio = second_row_numeric_count / total_cols
        
        # 1行目の70%以上が文字列で、2行目の30%以上が数値の場合
        return text_ratio >= 0.7 and numeric_ratio >= 0.3
    
    def _clean_cell_value(self, value: Any) -> Any:
        """
        セル値をクリーニング
        
        Args:
            value: セル値
            
        Returns:
            Any: クリーニング済みの値
        """
        if value is None:
            return None
        
        # 文字列の場合
        if isinstance(value, str):
            # 前後の空白を除去
            cleaned = value.strip()
            if cleaned == '':
                return None
            return cleaned
        
        # 数値の場合はそのまま返す
        if isinstance(value, (int, float)):
            return value
        
        # 日付時刻オブジェクトの場合は文字列に変換
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        
        # その他の場合は文字列に変換
        return str(value)
    
    def _resolve_duplicate_headers(self, headers: List[Any]) -> List[str]:
        """
        重複するヘッダー名を解決
        
        Args:
            headers: ヘッダーリスト
            
        Returns:
            List[str]: 重複を解決したヘッダーリスト
        """
        resolved_headers = []
        header_counts = {}
        
        for header in headers:
            # None や空文字の場合は自動生成
            if header is None or str(header).strip() == '':
                base_name = f"Column_{len(resolved_headers) + 1}"
            else:
                base_name = str(header).strip()
            
            # 重複チェック
            if base_name in header_counts:
                header_counts[base_name] += 1
                final_name = f"{base_name}_{header_counts[base_name]}"
            else:
                header_counts[base_name] = 0
                final_name = base_name
            
            resolved_headers.append(final_name)
        
        return resolved_headers
    
    def _empty_result(self, errors: List[str], warnings: List[str]) -> ExcelParseResult:
        """
        空の解析結果を生成
        
        Args:
            errors: エラーリスト
            warnings: 警告リスト
            
        Returns:
            ExcelParseResult: 空の解析結果
        """
        return ExcelParseResult(
            data={},
            sheet_names=[],
            row_counts={},
            headers={},
            has_headers={},
            errors=errors,
            warnings=warnings
        )